"""Phase 0.5 산출물(account_segmentation.json + accounts_master.json)을 사람이 바로 확인 가능한
엑셀 참고본으로 변환한다. 시트1 "공통대분류"(대분류/부서범위/등장부서/세부계정수/확정방식/비고),
시트2 "특정대분류"(부서별로 그룹핑, 대분류/세부계정수/확정방식/비고), 시트3 "세부계정"
(대분류/계정코드/계정명/적용부서/적용방식/계정 설명 — 계정코드 단위로 1행씩, 여러 부서에
걸쳐 등장해도 중복 없이 "적용부서" 컬럼에 목록으로 합쳐 표시). 읽기 전용 참고본이며,
실제 원가동인 확정 UI는 `/streamlit_app/app.py`(Streamlit 앱)가 담당한다.

`build_confirm_data()`는 이 파일과 Streamlit 앱이 공유하는 데이터 조립 함수다 —
account_segmentation.json + accounts_master.json을 읽어 대분류 목록(공통/특정 통합, `scope`
필드로 구분)과 각 대분류의 1~3순위 원가동인 추천, 세부계정별 원칙준용/예외지정 상태를
하나의 딕셔너리로 합친다. Streamlit 앱에서 이 모듈을 그대로 import해서 재사용한다
(화면단만 다르고 데이터 조립 로직은 하나로 유지).

Usage:
    python write_segmentation_excel.py <account_segmentation.json> <accounts_master.json> <output_xlsx>
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / "batch-tracker" / "scripts"))
from category_utils import resolve_account_effective  # noqa: E402

HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
GROUP_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HEADER_FONT = Font(bold=True)
GROUP_FONT = Font(bold=True)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_common_sheet(wb, seg):
    ws = wb.active
    ws.title = "공통대분류"
    headers = ["대분류", "부서범위", "등장부서", "세부계정수", "확정방식", "비고"]
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for cat in seg.get("common_categories", []):
        ws.append([
            cat["category"],
            cat.get("부서범위", ""),
            ", ".join(cat.get("departments", [])),
            len(cat.get("sub_accounts", [])),
            cat.get("확정방식", ""),
            cat.get("비고", ""),
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws, [18, 10, 30, 10, 12, 50])
    ws.freeze_panes = "A2"


def write_specific_sheet(wb, seg):
    ws = wb.create_sheet("특정대분류")
    headers = ["대분류", "세부계정수", "확정방식", "비고"]
    by_dept: dict[str, list] = {}
    for cat in seg.get("department_specific_categories", []):
        by_dept.setdefault(cat["department"], []).append(cat)

    row_idx = 1
    for dept in sorted(by_dept.keys()):
        ws.cell(row=row_idx, column=1, value=f"부서: {dept}")
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).fill = GROUP_FILL
            ws.cell(row=row_idx, column=col).font = GROUP_FONT
        row_idx += 1

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row_idx += 1

        for cat in by_dept[dept]:
            ws.cell(row=row_idx, column=1, value=cat["category"])
            ws.cell(row=row_idx, column=2, value=len(cat.get("sub_accounts", [])))
            ws.cell(row=row_idx, column=3, value=cat.get("확정방식", ""))
            ws.cell(row=row_idx, column=4, value=cat.get("비고", ""))
            row_idx += 1

        row_idx += 1  # 부서 간 빈 행

    for row in ws.iter_rows():
        for cell in row:
            if cell.alignment.wrap_text is None or not cell.alignment.wrap_text:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws, [18, 10, 12, 50])
    ws.freeze_panes = "A1"


def write_sub_account_sheet(wb, seg, master):
    ws = wb.create_sheet("세부계정")
    headers = ["대분류", "계정코드", "계정명", "적용부서", "적용방식", "계정 설명"]
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    all_categories = seg.get("common_categories", []) + seg.get("department_specific_categories", [])
    for cat in all_categories:
        for sa in cat.get("sub_accounts", []):
            eff = resolve_account_effective(master, sa["계정코드"])
            acc = master.get("accounts", {}).get(sa["계정코드"], {})
            desc = (acc.get("설명") or {}).get("내용", "")
            ws.append([
                cat["category"], sa["계정코드"], sa["계정명"],
                ", ".join(sa.get("등장부서", [])), eff["적용방식"], desc,
            ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws, [18, 12, 22, 24, 12, 40])
    ws.freeze_panes = "A2"


def _recommended_drivers(cat_state: dict) -> list:
    drivers = (cat_state.get("원가동인") or {}).get("recommended_drivers") or []
    # rank 오름차순 보장 + 필요한 키만 추출
    return [
        {"rank": d.get("rank"), "driver": d.get("driver", ""), "reason": d.get("reason", ""),
         "출처": d.get("근거출처", "")}
        for d in sorted(drivers, key=lambda d: d.get("rank") or 99)
    ]


def _merge_specific_entries_by_category(entries: list[dict]) -> list[dict]:
    """department_specific_categories는 부서 단위로 항목이 나뉘어 있다 — 원래 규칙 기반
    특정 대분류는 부서가 1개뿐이라 대분류당 항목도 1개였지만, Phase 0.5 재확인에서
    "특정전환"된 대분류는 여러 부서에 걸쳐 있어도 같은 대분류명으로 부서 수만큼 항목이
    생긴다(예: 17개 부서 x "보험료" = 17항목). accounts_master.json의 categories[대분류]는
    부서 무관 단일 레코드이므로, 화면에는 대분류명 기준으로 병합한 카드 하나만 보여준다 —
    그렇지 않으면 같은 카드가 부서 수만큼 중복 렌더링되어 Streamlit 위젯 키가 충돌한다
    (StreamlitDuplicateElementKey)."""
    merged: dict[str, dict] = {}
    for c in entries:
        cat_name = c["category"]
        entry = merged.setdefault(cat_name, {"category": cat_name, "departments": [], "sub_accounts": [], "_seen": set()})
        depts = c.get("departments") or ([c["department"]] if c.get("department") else [])
        for d in depts:
            if d not in entry["departments"]:
                entry["departments"].append(d)
        for sa in c.get("sub_accounts", []):
            if sa["계정코드"] not in entry["_seen"]:
                entry["_seen"].add(sa["계정코드"])
                entry["sub_accounts"].append(sa)
    for entry in merged.values():
        del entry["_seen"]
    return list(merged.values())


def build_confirm_data(seg: dict, master: dict) -> dict:
    categories_state = master.get("categories", {})

    accounts_state = master.get("accounts", {})

    categories = []
    for scope, seg_key in (("common", "common_categories"), ("specific", "department_specific_categories")):
        raw_entries = seg.get(seg_key, [])
        entries = _merge_specific_entries_by_category(raw_entries) if scope == "specific" else raw_entries
        for c in entries:
            cat_name = c["category"]
            cat_state = categories_state.get(cat_name, {})
            needs_review = bool(cat_state.get("추가판단필요여부"))

            sub_accounts = []
            for sa in c.get("sub_accounts", []):
                code = sa["계정코드"]
                eff = resolve_account_effective(master, code)
                acc = accounts_state.get(code, {})
                sub_accounts.append({
                    "code": code, "name": sa["계정명"], "departments": sa.get("등장부서", []),
                    "applyMode": eff["적용방식"],
                    "exception": eff if eff["적용방식"] == "예외지정" else None,
                    "description": acc.get("설명"),
                })

            categories.append({
                "category": cat_name,
                "scope": scope,
                "departments": c.get("departments") or [c.get("department")],
                "fourType": cat_state.get("four_type") or "추가판단 필요",
                "needsReview": needs_review,
                "drivers": [] if needs_review else _recommended_drivers(cat_state),
                "reason": cat_state.get("분류근거") or "",
                "subAccounts": sub_accounts,
            })

    return {
        "generated_at": seg.get("generated_at", ""),
        "departments": seg.get("departments_scanned", []),
        "categories": categories,
        "needs_llm_recheck": seg.get("needs_llm_recheck", []),
    }


def main():
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)

    seg_path, master_path, out_xlsx = sys.argv[1], sys.argv[2], sys.argv[3]

    with open(seg_path, encoding="utf-8") as f:
        seg = json.load(f)
    with open(master_path, encoding="utf-8") as f:
        master = json.load(f)

    if seg.get("needs_llm_recheck"):
        cats = ", ".join(r["대분류"] for r in seg["needs_llm_recheck"])
        print(f"경고: needs_llm_recheck가 비어있지 않습니다 ({cats}) — Phase 0.5 미완료 상태로 산출물을 생성합니다.")

    wb = Workbook()
    write_common_sheet(wb, seg)
    write_specific_sheet(wb, seg)
    write_sub_account_sheet(wb, seg, master)

    Path(out_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)

    print(
        f"공통 대분류 {len(seg.get('common_categories', []))}건, "
        f"특정 대분류 {len(seg.get('department_specific_categories', []))}건 -> {out_xlsx}"
    )


if __name__ == "__main__":
    main()
