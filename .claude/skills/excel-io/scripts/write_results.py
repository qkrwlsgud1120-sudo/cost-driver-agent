"""검증 완료된 배치 JSON을 최종 원가동인 엑셀로 저장한다.

export_mode:
  - editable (기본): 회계사가 엑셀에서 직접 확정·재업로드할 수 있는 편집용 시트
  - report: Streamlit 「전체 저장」 후 산출하는 확정 결과 보고서 (4-type·확정·근거·비고)

Usage:
    python write_results.py <validated_json> <output_xlsx> [--mode editable|report]
"""
import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAX_RANKS = 3
KST = timezone(timedelta(hours=9))

COLUMNS_EDITABLE = [
    "대분류", "계정코드", "계정명", "계정 설명", "부서", "4-type 분류", "판단 경로",
    "추천 1순위", "근거 1순위", "추천 2순위", "근거 2순위", "추천 3순위", "근거 3순위",
    "4-type 분류(사람 확정)", "확정 순위", "원가동인 적용 방식", "확정 여부", "확정 원가동인", "비고",
]
# read_confirmations.py 하위 호환 — 구 헤더명 매핑
COLUMNS_LEGACY = [
    "대분류", "계정코드", "계정명", "계정 설명", "부서", "4-type 분류", "판단 경로",
    "AI 추천 1순위", "근거 1순위", "AI 추천 2순위", "근거 2순위", "AI 추천 3순위", "근거 3순위",
    "4-type 분류(사람 확정)", "확정 순위", "원가동인 적용 방식", "확정 여부", "확정 원가동인", "비고",
]

REPORT_COLUMNS = [
    "대분류", "계정코드", "계정명", "계정 설명", "부서", "4-type 분류", "판단 경로", "원가동인 적용 방식",
    "확정 여부", "확정 원가동인", "확정 근거", "비고",
]

FOUR_TYPES = ["직접귀속형", "배부형", "공통비형", "기타"]

REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
HEADER_FONT = Font(bold=True)
EDITABLE_FILL = PatternFill(start_color="EAF7EA", end_color="EAF7EA", fill_type="solid")

REPORT_TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
REPORT_TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
REPORT_SUBTITLE_FONT = Font(color="5A6B7D", size=10)
REPORT_HEADER_FILL = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
REPORT_HEADER_FONT = Font(bold=True, color="1F4E79", size=10)
REPORT_ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
REPORT_STATUS_FILLS = {
    "승인": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "확정": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "수정": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "미확정": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
}
REPORT_AMBIGUOUS_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)

CONFIRM_STATUS_OPTIONS_NORMAL = '"미확정,승인,수정"'
CONFIRM_STATUS_OPTIONS_AMBIGUOUS = '"미확정,확정"'
HUMAN_FOUR_TYPE_OPTIONS = f'"{",".join(FOUR_TYPES)}"'
LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)


def rank_options_formula(n_ranks: int) -> str:
    opts = [f"{i}순위" for i in range(1, n_ranks + 1)] + ["직접입력"]
    return '"' + ",".join(opts) + '"'


def _driver_rank_map(r: dict) -> dict[int, dict]:
    ranks: dict[int, dict] = {}
    for item in r.get("recommended_drivers", []) or []:
        rank = item.get("rank")
        if rank and 1 <= rank <= MAX_RANKS:
            ranks[rank] = {
                "driver": item.get("driver", ""),
                "reason": item.get("reason", ""),
            }
    return ranks


def resolve_four_type_display(r: dict) -> str:
    if r.get("four_type"):
        return r["four_type"]
    if r.get("사람확정four_type"):
        return r["사람확정four_type"]
    if r.get("추가판단필요여부"):
        return "추가판단 필요"
    return ""


def resolve_confirm_reason(r: dict) -> str:
    status = r.get("확정여부") or "미확정"
    rank = r.get("확정순위")
    ranks = _driver_rank_map(r)

    if status in ("", "미확정"):
        if r.get("추가판단필요여부"):
            return r.get("분류근거") or ""
        return ""

    if status == "수정":
        return "회계사 직접 입력으로 확정"

    if status == "확정":
        if ranks:
            first = sorted(ranks.items())[0][1]
            return first.get("reason") or "회계사 직접 확정"
        return r.get("분류근거") or "회계사 직접 확정"

    if status == "승인":
        pick = rank if rank in (1, 2, 3) else 1
        if pick in ranks and ranks[pick].get("reason"):
            return ranks[pick]["reason"]
        if ranks:
            first = sorted(ranks.items())[0][1]
            return first.get("reason") or ""
        return r.get("분류근거") or ""

    return r.get("분류근거") or ""


def resolve_note(r: dict) -> str:
    if r.get("검토필요"):
        return "검증 단계에서 재확인 필요로 표시됨"
    status = r.get("확정여부") or "미확정"
    if r.get("추가판단필요여부") and status == "미확정":
        return "4-type·원가동인 추가판단 필요"
    if status == "미확정":
        return "확정 대기"
    if status == "승인" and r.get("확정순위") in (2, 3):
        return f"{r['확정순위']}순위 추천안 확정"
    return ""


def to_report_row(r: dict) -> dict:
    status = r.get("확정여부") or "미확정"
    driver = r.get("확정원가동인") or ""
    if status == "승인" and not driver:
        ranks = _driver_rank_map(r)
        pick = r.get("확정순위") if r.get("확정순위") in (1, 2, 3) else 1
        driver = (ranks.get(pick) or {}).get("driver") or (ranks.get(1) or {}).get("driver") or ""

    return {
        "대분류": r.get("대분류", ""),
        "계정코드": r["계정코드"],
        "계정명": r["계정명"],
        "계정 설명": r.get("계정설명", ""),
        "부서": r["부서명"],
        "4-type 분류": resolve_four_type_display(r),
        "판단 경로": r.get("판단경로") or "확인불가",
        "원가동인 적용 방식": r.get("적용방식", "원칙준용"),
        "확정 여부": status,
        "확정 원가동인": driver,
        "확정 근거": resolve_confirm_reason(r),
        "비고": resolve_note(r),
        "_추가판단필요": bool(r.get("추가판단필요여부")),
        "_status": status,
    }


def to_row(r: dict) -> dict:
    needs_review = bool(r.get("추가판단필요여부", False))
    need_review_flag = needs_review or bool(r.get("검토필요", False))
    ranks = _driver_rank_map(r)
    n_ranks = len(ranks)

    if needs_review:
        four_type_display = "추가판단 필요"
        reason_fallback = r.get("분류근거", "") or ""
        note = (
            "4-type 미확정 — '4-type 분류(사람 확정)'만 채우면 원가동인 추천 재실행 대상. "
            "4-type과 원가동인을 함께 채우고 '확정'을 선택하면 직접 확정 반영."
        )
    else:
        four_type_display = r.get("four_type", "")
        note = "검증 단계에서 재확인 필요" if r.get("검토필요") else ""

    prefill_status = r.get("확정여부")
    prefill_driver = r.get("확정원가동인")
    prefill_rank = r.get("확정순위")
    prefill_rank_display = (
        f"{prefill_rank}순위" if prefill_rank in (1, 2, 3)
        else ("직접입력" if prefill_status == "수정" else "")
    )

    row = {
        "대분류": r.get("대분류", ""),
        "계정코드": r["계정코드"],
        "계정명": r["계정명"],
        "계정 설명": r.get("계정설명", ""),
        "부서": r["부서명"],
        "4-type 분류": four_type_display,
        "판단 경로": r.get("판단경로") or "확인불가",
        "4-type 분류(사람 확정)": (r.get("사람확정four_type") or "") if needs_review else "",
        "확정 순위": prefill_rank_display if not needs_review else "",
        "원가동인 적용 방식": r.get("적용방식", "원칙준용"),
        "확정 여부": prefill_status or "미확정",
        "확정 원가동인": (prefill_driver or "") if needs_review else (prefill_driver if prefill_driver else None),
        "비고": note,
        "_추가판단필요": needs_review,
        "_검토필요": need_review_flag,
        "_n_ranks": n_ranks,
    }
    for i in range(1, MAX_RANKS + 1):
        item = ranks.get(i)
        row[f"추천 {i}순위"] = item["driver"] if item else ""
        row[f"근거 {i}순위"] = item["reason"] if item else ""
    if needs_review:
        row["근거 1순위"] = reason_fallback
    return row


def style_editable_header(ws):
    col = {name: i + 1 for i, name in enumerate(COLUMNS_EDITABLE)}
    for col_idx, name in enumerate(COLUMNS_EDITABLE, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.protection = LOCKED
    ws.freeze_panes = "A2"
    widths = [18, 12, 22, 30, 12, 12, 14, 20, 40, 20, 40, 20, 40, 16, 10, 14, 10, 24, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return col


def append_editable_rows(ws, rows: list[dict], start_row: int, col: dict):
    status_col_letter = get_column_letter(col["확정 여부"])
    rank_col_letter = get_column_letter(col["확정 순위"])
    driver_col_letters = {i: get_column_letter(col[f"추천 {i}순위"]) for i in range(1, MAX_RANKS + 1)}

    ambiguous_row_nums = []
    normal_row_nums_by_n_ranks: dict[int, list[int]] = {}

    for offset, row in enumerate(rows):
        r = start_row + offset
        ambiguous = row["_추가판단필요"]
        if ambiguous:
            ambiguous_row_nums.append(r)
        else:
            normal_row_nums_by_n_ranks.setdefault(row["_n_ranks"], []).append(r)

        for name in COLUMNS_EDITABLE:
            cell = ws.cell(row=r, column=col[name])
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if name == "확정 원가동인" and not ambiguous and row[name] is None:
                nested = '""'
                for i in range(MAX_RANKS, 0, -1):
                    nested = f'IF({rank_col_letter}{r}="{i}순위",{driver_col_letters[i]}{r},{nested})'
                cell.value = f'=IF({status_col_letter}{r}="승인",{nested},"")'
            else:
                cell.value = row[name] if row[name] is not None else ""

            if name == "확정 순위":
                editable = not ambiguous
            elif name == "4-type 분류(사람 확정)":
                editable = ambiguous
            elif name in ("확정 여부", "확정 원가동인"):
                editable = True
            else:
                editable = False

            if editable:
                cell.protection = UNLOCKED
                cell.fill = EDITABLE_FILL
            else:
                cell.protection = LOCKED

        if row["_검토필요"]:
            for name in COLUMNS_EDITABLE:
                ws.cell(row=r, column=col[name]).fill = REVIEW_FILL
            for name in ("확정 여부", "확정 원가동인"):
                ws.cell(row=r, column=col[name]).fill = EDITABLE_FILL
            if not ambiguous:
                ws.cell(row=r, column=col["확정 순위"]).fill = EDITABLE_FILL
            if ambiguous:
                ws.cell(row=r, column=col["4-type 분류(사람 확정)"]).fill = EDITABLE_FILL

    def add_list_validation(col_letter, formula, row_nums, allow_blank=False):
        if not row_nums:
            return
        dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank, showDropDown=False)
        dv.error = "목록에 있는 값만 선택하세요."
        dv.errorTitle = "잘못된 값"
        ws.add_data_validation(dv)
        for rn in row_nums:
            dv.add(f"{col_letter}{rn}")

    normal_row_nums_all = [rn for rows_ in normal_row_nums_by_n_ranks.values() for rn in rows_]
    add_list_validation(status_col_letter, CONFIRM_STATUS_OPTIONS_NORMAL, normal_row_nums_all)
    add_list_validation(status_col_letter, CONFIRM_STATUS_OPTIONS_AMBIGUOUS, ambiguous_row_nums)
    add_list_validation(
        get_column_letter(col["4-type 분류(사람 확정)"]), HUMAN_FOUR_TYPE_OPTIONS,
        ambiguous_row_nums, allow_blank=True,
    )
    for n_ranks, row_nums in normal_row_nums_by_n_ranks.items():
        add_list_validation(rank_col_letter, rank_options_formula(max(n_ranks, 1)), row_nums)


def _header_matches(existing: list, expected: list) -> bool:
    return existing == expected or existing == COLUMNS_LEGACY


def write_report_workbook(
    records: list[dict], output_xlsx, overwrite: bool = True, dept_label: str | None = None,
) -> int:
    """확정 결과 보고서 — 4-type·확정여부·확정 원가동인·근거·비고만 표시.

    dept_label을 지정하면 서브타이틀에 그 값을 쓴다(여러 부서 레코드를 한 워크북에
    합칠 때 "전체" 등으로 표시하기 위함). 지정하지 않으면 기존처럼 첫 레코드의
    부서명을 그대로 쓴다(단일 부서 리포트, 하위 호환).
    """
    rows = [to_report_row(r) for r in records]
    out_path = Path(output_xlsx)
    if overwrite and out_path.exists():
        out_path.unlink()

    dept = dept_label if dept_label is not None else (records[0]["부서명"] if records else "")
    generated = datetime.now(KST).strftime("%Y-%m-%d %H:%M")
    n_cols = len(REPORT_COLUMNS)
    last_col = get_column_letter(n_cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "원가동인 확정"

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = "원가동인 확정 결과"
    title_cell.font = REPORT_TITLE_FONT
    title_cell.fill = REPORT_TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    sub = ws["A2"]
    sub.value = f"부서: {dept}  ·  생성: {generated}  ·  계정 {len(rows)}건"
    sub.font = REPORT_SUBTITLE_FONT
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 4
    widths = [18, 12, 24, 30, 12, 12, 14, 14, 10, 22, 48, 28]
    for col_idx, name in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.font = REPORT_HEADER_FONT
        cell.fill = REPORT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    data_start = header_row + 1
    for offset, row in enumerate(rows):
        r = data_start + offset
        if row["_추가판단필요"] and row["_status"] == "미확정":
            row_fill = REPORT_AMBIGUOUS_FILL
        else:
            row_fill = REPORT_STATUS_FILLS.get(row["_status"])
            if not row_fill and offset % 2 == 1:
                row_fill = REPORT_ALT_FILL

        for col_idx, name in enumerate(REPORT_COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=row.get(name, ""))
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center" if name in ("대분류", "계정코드", "부서", "4-type 분류", "판단 경로", "원가동인 적용 방식", "확정 여부") else "left",
            )
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

    ws.sheet_view.showGridLines = False
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(rows)


def write_workbook(
    records: list[dict],
    output_xlsx,
    overwrite: bool = False,
    export_mode: str = "editable",
    dept_label: str | None = None,
) -> int:
    """records를 output_xlsx에 씬다.

    export_mode:
      - report: 확정 결과 보고서 (Streamlit 저장용)
      - editable: 회계사 편집·재업로드용 (Phase 1 배치 산출)

    dept_label은 report 모드에서만 쓰인다(write_report_workbook 참조) — 여러 부서
    레코드를 한 워크북으로 합칠 때 서브타이틀을 "전체" 등으로 지정하기 위함.
    """
    if export_mode == "report":
        return write_report_workbook(records, output_xlsx, overwrite=overwrite or True, dept_label=dept_label)

    rows = [to_row(r) for r in records]
    out_path = Path(output_xlsx)

    if overwrite and out_path.exists():
        out_path.unlink()

    if out_path.exists():
        wb = load_workbook(out_path)
        ws = wb["원가동인추천"]
        existing_header = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMNS_EDITABLE) + 1)]
        if not _header_matches(existing_header, COLUMNS_EDITABLE):
            raise ValueError(
                f"{out_path}의 기존 컬럼 구조가 다릅니다 (기존: {existing_header}). "
                "구버전 산출물이면 새 파일 경로로 다시 생성하세요."
            )
        start_row = ws.max_row + 1
        col = {name: i + 1 for i, name in enumerate(COLUMNS_EDITABLE)}
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "원가동인추천"
        col = style_editable_header(ws)
        start_row = 2

    append_editable_rows(ws, rows, start_row, col)

    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return start_row - 2 + len(rows)


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    validated_json, output_xlsx = sys.argv[1], sys.argv[2]
    export_mode = "editable"
    if len(sys.argv) >= 4 and sys.argv[3] == "--mode":
        export_mode = sys.argv[4] if len(sys.argv) >= 5 else "editable"

    with open(validated_json, encoding="utf-8") as f:
        records = json.load(f)

    total_rows = write_workbook(records, output_xlsx, export_mode=export_mode)
    label = "확정 보고서" if export_mode == "report" else "편집용 시트"
    print(f"{len(records)}건 → {output_xlsx} (총 {total_rows}건, {label})")


if __name__ == "__main__":
    main()
