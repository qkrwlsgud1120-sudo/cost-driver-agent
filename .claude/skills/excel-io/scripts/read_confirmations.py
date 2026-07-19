"""회계사가 편집해 재업로드한 batch_{id}-{부서}_final.xlsx(editable 모드)를 읽어 원본
행 값을 그대로 반환한다 — 대분류 스키마(`write_results.py`의 COLUMNS_EDITABLE) 기준.

**"새로 기입된 항목인지" 판정은 이 스크립트의 책임이 아니다.** 엑셀 파일 자체에는
"이전에 무엇이었는지"에 대한 정보가 없고, 그건 accounts_master.json(현재 상태)과
대조해야만 알 수 있다 — 그 비교와 반영은 `track_batch.py confirm()`이 수행한다.
이 스크립트는 순수하게 엑셀 → 행 단위 JSON 변환만 한다.

Usage:
    python read_confirmations.py <accountant_edited_final.xlsx> <output_rows.json>
"""
import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import load_workbook

KST = timezone(timedelta(hours=9))

# write_results.py: COLUMNS_EDITABLE과 정확히 일치해야 한다(대분류 도입 이후 스키마).
COLUMNS = [
    "대분류", "계정코드", "계정명", "계정 설명", "부서", "4-type 분류", "판단 경로",
    "추천 1순위", "근거 1순위", "추천 2순위", "근거 2순위", "추천 3순위", "근거 3순위",
    "4-type 분류(사람 확정)", "확정 순위", "원가동인 적용 방식", "확정 여부", "확정 원가동인", "비고",
]
# 구 헤더명(계정코드 기반 스키마 시절 — write_results.py: COLUMNS_LEGACY)과의 호환.
COLUMNS_LEGACY = [
    "대분류", "계정코드", "계정명", "계정 설명", "부서", "4-type 분류", "판단 경로",
    "AI 추천 1순위", "근거 1순위", "AI 추천 2순위", "근거 2순위", "AI 추천 3순위", "근거 3순위",
    "4-type 분류(사람 확정)", "확정 순위", "원가동인 적용 방식", "확정 여부", "확정 원가동인", "비고",
]

COL_INDEX = {name: i + 1 for i, name in enumerate(COLUMNS)}


def read_confirmations(xlsx_path: str) -> dict:
    wb_values = load_workbook(xlsx_path, data_only=True)
    ws = wb_values["원가동인추천"]

    header = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMNS) + 1)]
    if header != COLUMNS and header != COLUMNS_LEGACY:
        raise ValueError(f"컬럼 구조가 예상과 다릅니다 (파일: {xlsx_path}, 헤더: {header})")

    def cell(row: int, name: str):
        return ws.cell(row=row, column=COL_INDEX[name]).value

    rows = []
    for row in range(2, ws.max_row + 1):
        code = cell(row, "계정코드")
        if code is None:
            continue

        four_type_display = cell(row, "4-type 분류")
        rows.append({
            "대분류": cell(row, "대분류"),
            "계정코드": code,
            "계정명": cell(row, "계정명"),
            "부서": cell(row, "부서"),
            "needs_review": four_type_display == "추가판단 필요",
            "AI_4type": four_type_display,
            "판단경로": cell(row, "판단 경로"),
            "AI_1순위_동인": cell(row, "추천 1순위"),
            "사람확정4type": cell(row, "4-type 분류(사람 확정)"),
            "확정순위표시": cell(row, "확정 순위"),
            "원가동인_적용방식_표시": cell(row, "원가동인 적용 방식"),
            "확정여부": cell(row, "확정 여부"),
            "확정원가동인": cell(row, "확정 원가동인"),
            "비고": cell(row, "비고"),
        })

    return {
        "generated_at": datetime.now(KST).isoformat(),
        "source_file": str(xlsx_path),
        "rows": rows,
    }


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    xlsx_path, out_path = sys.argv[1], sys.argv[2]
    result = read_confirmations(xlsx_path)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"{len(result['rows'])}행 파싱 -> {out_path}")


if __name__ == "__main__":
    main()
